"""Chesterfield County ACA bulk Code Violation report puller.

Pulls the date-range Code Violation report from the Chesterfield Accela
Citizen Access portal. Anonymous access, no login required.

URL contract (verified 2026-05-25, see test_chesterfield_aca_recon.py):
    Form:   https://aca-prod.accela.com/CHESTERFIELD/Report/ReportParameter.aspx
            ?module=&reportID=9735&reportType=LINK_REPORT_LIST
    Inputs: #Date_11907 (Start), #Date_11908 (End)  — MaskedEdit MM/DD/YYYY
    Submit: #btnSave (the visible "Submit" link)
    Result: opens new tab as XLSX download (CE_Code Violation_V2.xlsx)

XLSX schema (header at row 3, 0-indexed):
    Record Type | Record ID | Submittal Date | Record Status Date
                | Record Status | Code Section | Property Address

Volume: ~370 records per 30-day window in May 2026 recon (~12/day).

Diff strategy: dedup by Record ID. State file `chesterfield_aca_state.json`
at PROJECT_ROOT holds the set of previously-seen Record IDs.

See memory: chesterfield-aca-code-violation-report for the operator's
discovery context (newly-launched public report, May 2026).
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import logging
import re
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import config
from notice_parser import NoticeData

logger = logging.getLogger(__name__)


# ── Config ────────────────────────────────────────────────────────────

CHESTERFIELD_ACA_STATE_FILE: Path = config.PROJECT_ROOT / "chesterfield_aca_state.json"
CHESTERFIELD_ACA_STATE_SCHEMA_VERSION: int = 1

REPORT_URL: str = (
    "https://aca-prod.accela.com/CHESTERFIELD/Report/ReportParameter.aspx"
    "?module=&reportID=9735&reportType=LINK_REPORT_LIST"
)
LANDING_URL: str = "https://aca-prod.accela.com/CHESTERFIELD/Default.aspx"

# Selectors verified in recon.
SEL_START_DATE: str = "#Date_11907"
SEL_END_DATE: str = "#Date_11908"
SEL_SUBMIT: str = "#btnSave"

# Default date window for the first-ever pull. After that the state file
# tracks the last fetched range and we only fetch yesterday-to-today.
DEFAULT_FIRST_PULL_DAYS: int = 90

# Cap records per single pull — if a window returns more than this, something
# is wrong (or the operator picked a too-wide range) and we abort.
MAX_RECORDS_PER_PULL: int = 5000

# Code Section allow-list — substring match (case-insensitive) on the
# "Code Section" column. ONLY records whose case cites at least one of these
# signals get emitted as NoticeData (and thus uploaded to DataSift).
#
# Default = "vacant" — catches `11-32(b) - Tall Grass/Vacant` and any future
# Vacant-named sections. Operator confirmed 2026-05-25: vacant-tagged
# properties are highest motivation.
#
# Candidate extensions (from May 2026 30-day recon, 367 rows):
#   "Discarded Materials"   — 41 records, abandoned-property signal
#   "Inoperable Vehicle"    — 39 records, dead car on property
#   "Trash/Litter"          — 7 records, neglect
#   "Tall Grass/Occupied"   — 132 records, weaker (owner present)
#
# All records that DON'T match the filter still get tracked in
# `chesterfield_aca_state.json` for dedup, so widening the filter later
# requires deleting state to re-emit historical hits. Use `--aca-all-violations`
# (CLI) or `aca_all_violations=true` (Apify input) to bypass the filter.
HIGH_MOTIVATION_CODE_SECTIONS: list[str] = [
    "vacant",
]


def _is_high_motivation(code_section: str) -> bool:
    """Return True if a code section signals a high-motivation property."""
    if not code_section:
        return False
    lowered = code_section.lower()
    return any(pat.lower() in lowered for pat in HIGH_MOTIVATION_CODE_SECTIONS)


# ── Data types ────────────────────────────────────────────────────────


@dataclass
class ChesterfieldCaseRow:
    """One row from the ACA Code Violation XLSX export."""
    record_type: str        # "Property Maintenance" | "Zoning Code Compliance"
    record_id: str          # e.g. "PM26-0029" | "CE26-0877"
    submittal_date: str     # YYYY-MM-DD
    status_date: str        # YYYY-MM-DD
    status: str             # e.g. "In Violation"
    code_section: str       # e.g. "304.12 - Handrails and guards"
    property_address: str   # e.g. "219 HAZELMERE DR" (uppercase, no city/zip)


# ── XLSX parsing ──────────────────────────────────────────────────────


def parse_report_xlsx(xlsx_path: Path) -> list[ChesterfieldCaseRow]:
    """Parse the Chesterfield ACA Code Violation report.

    Header is at row 3 (0-indexed). Earlier rows are title + blanks.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl required for Chesterfield ACA report parsing — "
            "should already be in requirements.txt"
        ) from exc

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    logger.info(
        "Loaded XLSX %s (sheet=%s, %d rows, %d cols)",
        xlsx_path.name, ws.title, ws.max_row, ws.max_column,
    )

    # Find header row — first row with >=3 non-empty cells matching schema.
    header_row_idx = None
    col_map: dict[str, int] = {}
    expected = {
        "Record Type", "Record ID", "Submittal Date",
        "Record Status Date", "Record Status", "Code Section", "Property Address",
    }
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=0):
        present = {str(c).strip(): i for i, c in enumerate(row) if c}
        overlap = expected & set(present.keys())
        if len(overlap) >= 5:
            header_row_idx = idx
            col_map = {k: present[k] for k in expected if k in present}
            break

    if header_row_idx is None:
        raise RuntimeError("Header row not found in XLSX")
    logger.info("Header row %d, columns: %s", header_row_idx, col_map)

    rows: list[ChesterfieldCaseRow] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=0):
        if r_idx <= header_row_idx:
            continue
        try:
            rec = _row_to_record(row, col_map)
        except Exception as e:
            logger.debug("Row %d skipped: %s", r_idx, e)
            continue
        if rec is None:
            continue
        rows.append(rec)

    return rows


def _row_to_record(row: tuple, col_map: dict[str, int]) -> ChesterfieldCaseRow | None:
    """Convert one XLSX row tuple to a ChesterfieldCaseRow."""
    def _get(key: str) -> str:
        i = col_map.get(key)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    record_id = _get("Record ID")
    address = _get("Property Address")
    if not record_id or not address:
        return None

    return ChesterfieldCaseRow(
        record_type=_get("Record Type"),
        record_id=record_id,
        submittal_date=_get("Submittal Date"),
        status_date=_get("Record Status Date"),
        status=_get("Record Status"),
        code_section=_get("Code Section"),
        property_address=address,
    )


# ── Playwright fetch ──────────────────────────────────────────────────


async def fetch_report_xlsx(
    start_date: date,
    end_date: date,
    *,
    headless: bool = True,
    slow_mo_ms: int = 0,
) -> Path:
    """Drive the ACA report flow and return the path to the downloaded XLSX.

    Caller owns the returned file (temp dir under output/).
    """
    from playwright.async_api import async_playwright, TimeoutError as PwTimeout

    out_dir = Path(tempfile.mkdtemp(
        prefix="chesterfield_aca_",
        dir=str(config.OUTPUT_DIR),
    ))
    start_str = start_date.strftime("%m/%d/%Y")
    end_str = end_date.strftime("%m/%d/%Y")
    logger.info("Fetching ACA report %s → %s", start_str, end_str)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=headless,
            slow_mo=slow_mo_ms,
        )
        context = await browser.new_context(
            accept_downloads=True,
            viewport={"width": 1400, "height": 900},
        )
        page = await context.new_page()

        # Establish session (some ACA installs gate the report URL on a prior
        # landing-page visit). Cheap insurance.
        await page.goto(LANDING_URL, wait_until="domcontentloaded")
        await page.wait_for_timeout(2000)

        await page.goto(REPORT_URL, wait_until="domcontentloaded")
        await page.wait_for_timeout(2500)

        # MaskedEdit inputs: click → clear → type with per-keystroke delay,
        # then Tab to blur so the validator passes.
        for selector, value in ((SEL_START_DATE, start_str), (SEL_END_DATE, end_str)):
            await page.click(selector, timeout=10000)
            await page.keyboard.press("Control+A")
            await page.keyboard.press("Delete")
            await page.type(selector, value, delay=40)
            await page.keyboard.press("Tab")

        # Let the AjaxControlToolkit validators finish before submit; otherwise
        # the submit handler bails on a stale validation state.
        await page.wait_for_timeout(1200)

        # Submit triggers a SAME-PAGE XLSX download (no popup — verified in recon).
        # Server-side report generation can be slow (the report is built per-request
        # from a SQL query), so allow up to 3 minutes for the download to fire.
        try:
            async with page.expect_download(timeout=180000) as dl_info:
                await page.click(SEL_SUBMIT)
            download = await dl_info.value
        except PwTimeout as exc:
            await page.screenshot(path=str(out_dir / "submit_failed.png"), full_page=True)
            raise RuntimeError(
                "ACA report XLSX download did not start within 180s — "
                "check submit_failed.png"
            ) from exc

        target = out_dir / (download.suggested_filename or "report.xlsx")
        await download.save_as(str(target))
        logger.info("Downloaded ACA report: %s (%d bytes)", target, target.stat().st_size)

        await browser.close()
        return target


def fetch_report_xlsx_sync(
    start_date: date,
    end_date: date,
    *,
    headless: bool = True,
    slow_mo_ms: int = 0,
) -> Path:
    """Synchronous wrapper around the Playwright fetch."""
    return asyncio.run(
        fetch_report_xlsx(start_date, end_date, headless=headless, slow_mo_ms=slow_mo_ms)
    )


# ── State management ──────────────────────────────────────────────────


def load_state() -> dict:
    """Load puller state, returning a fresh shell if missing/empty."""
    state = config.load_state(CHESTERFIELD_ACA_STATE_FILE)
    if not state:
        return {
            "schema_version": CHESTERFIELD_ACA_STATE_SCHEMA_VERSION,
            "last_fetch_at": "",
            "last_window_start": "",
            "last_window_end": "",
            "known_record_ids": [],
        }
    state.setdefault("known_record_ids", [])
    return state


def save_state(state: dict) -> None:
    config.save_state(CHESTERFIELD_ACA_STATE_FILE, state)


# ── Top-level pull ────────────────────────────────────────────────────


def _to_notice(row: ChesterfieldCaseRow, today_iso: str) -> NoticeData:
    """Convert a parsed XLSX row into a SiftStack NoticeData."""
    # Use submittal_date if present (when the case was opened), fallback to today.
    date_added = row.submittal_date or today_iso

    # Build a concise raw_text capturing the key context for downstream consumers.
    raw_text = (
        f"Chesterfield County code enforcement case\n"
        f"Record Type: {row.record_type}\n"
        f"Record ID: {row.record_id}\n"
        f"Code Section: {row.code_section}\n"
        f"Status: {row.status}\n"
        f"Status Date: {row.status_date}\n"
        f"Submittal Date: {row.submittal_date}"
    )

    return NoticeData(
        date_added=date_added,
        address=row.property_address,
        city="",                    # XLSX has no city — Smarty fills it downstream
        state="VA",
        zip="",
        owner_name="",              # XLSX has no owner — assessor lookup downstream
        notice_type="code_violation",
        county="Chesterfield County",
        source_url=f"chesterfield_aca_report://{row.record_id}",
        raw_text=raw_text,
    )


def pull_new_records(
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    first_pull_days: int = DEFAULT_FIRST_PULL_DAYS,
    headless: bool = True,
    all_violations: bool = False,
) -> list[NoticeData]:
    """Fetch the ACA Code Violation report, diff against state, return new records.

    Window selection:
      - start_date/end_date overrides take priority
      - Otherwise: if state has never been written, use `first_pull_days` back from today
      - Otherwise: use the last window's end_date as the new start_date, end_date=today

    Filtering:
      By default, only records whose Code Section matches HIGH_MOTIVATION_CODE_SECTIONS
      are emitted (records not matching are still added to state for dedup).
      Pass all_violations=True to bypass the filter and emit every record.
    """
    state = load_state()
    today = datetime.now(timezone.utc).date()

    if end_date is None:
        end_date = today

    if start_date is None:
        last_end = state.get("last_window_end") or ""
        if last_end:
            try:
                last_end_d = datetime.strptime(last_end, "%Y-%m-%d").date()
                # Step back one day to overlap and re-catch any late-arriving cases
                start_date = last_end_d - timedelta(days=1)
            except ValueError:
                start_date = today - timedelta(days=first_pull_days)
        else:
            start_date = today - timedelta(days=first_pull_days)

    if start_date > end_date:
        logger.info("start_date > end_date — nothing to pull")
        return []

    # Fetch + parse
    xlsx_path = fetch_report_xlsx_sync(start_date, end_date, headless=headless)
    rows = parse_report_xlsx(xlsx_path)
    logger.info("Parsed %d rows from %s", len(rows), xlsx_path.name)

    if len(rows) > MAX_RECORDS_PER_PULL:
        raise RuntimeError(
            f"ACA report returned {len(rows)} rows — exceeds MAX_RECORDS_PER_PULL "
            f"({MAX_RECORDS_PER_PULL}). Refusing to update state. Reduce the window."
        )

    # Diff. A single Record ID can cite multiple Code Sections at the same
    # property, so we group rows by record_id and evaluate the filter against
    # the full set of code sections for that case. Records that don't match
    # are still added to state (dedup) but not emitted.
    known: set[str] = set(state.get("known_record_ids", []))
    today_iso = today.strftime("%Y-%m-%d")
    new_notices: list[NoticeData] = []
    seen_this_run: set[str] = set()
    emitted_count = 0
    filtered_count = 0

    # Group all rows in this XLSX by Record ID, preserving first-seen row order.
    grouped: dict[str, list[ChesterfieldCaseRow]] = {}
    for row in rows:
        grouped.setdefault(row.record_id, []).append(row)

    for record_id, group_rows in grouped.items():
        if record_id in known:
            continue
        seen_this_run.add(record_id)

        all_sections = [r.code_section for r in group_rows if r.code_section]
        matches_filter = all_violations or any(
            _is_high_motivation(s) for s in all_sections
        )
        if not matches_filter:
            filtered_count += 1
            continue

        # Emit one NoticeData per case. Use the first row as primary and
        # collapse all Code Sections into a pipe-joined raw_text line so the
        # full violation picture is preserved.
        primary = group_rows[0]
        if len(all_sections) > 1:
            joined = " | ".join(all_sections)
            primary = ChesterfieldCaseRow(
                record_type=primary.record_type,
                record_id=primary.record_id,
                submittal_date=primary.submittal_date,
                status_date=primary.status_date,
                status=primary.status,
                code_section=joined,
                property_address=primary.property_address,
            )
        new_notices.append(_to_notice(primary, today_iso))
        emitted_count += 1

    # Persist state — include ALL seen IDs, even filtered-out ones, so future
    # runs don't re-evaluate them. Operator can delete state to re-emit if the
    # filter is widened later.
    if seen_this_run:
        state["known_record_ids"] = sorted(known | seen_this_run)
    state["last_fetch_at"] = datetime.now(timezone.utc).isoformat()
    state["last_window_start"] = start_date.strftime("%Y-%m-%d")
    state["last_window_end"] = end_date.strftime("%Y-%m-%d")
    save_state(state)

    logger.info(
        "Chesterfield ACA delta: %d emitted / %d filtered-out / %d total new unique cases "
        "(%d raw rows parsed, %d total tracked in state, filter=%s)",
        emitted_count, filtered_count, len(seen_this_run),
        len(rows), len(state["known_record_ids"]),
        "ALL" if all_violations else f"high-motivation:{HIGH_MOTIVATION_CODE_SECTIONS}",
    )
    return new_notices


# ── Diagnostic CLI ────────────────────────────────────────────────────


def _cli() -> None:
    p = argparse.ArgumentParser(description="Chesterfield ACA Code Violation puller")
    p.add_argument("--start", help="Override start date YYYY-MM-DD")
    p.add_argument("--end", help="Override end date YYYY-MM-DD")
    p.add_argument("--first-pull-days", type=int, default=DEFAULT_FIRST_PULL_DAYS)
    p.add_argument("--headed", action="store_true", help="Run with visible browser")
    p.add_argument(
        "--all-violations",
        action="store_true",
        help="Bypass the HIGH_MOTIVATION_CODE_SECTIONS filter — emit every case",
    )
    args = p.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    start = datetime.strptime(args.start, "%Y-%m-%d").date() if args.start else None
    end = datetime.strptime(args.end, "%Y-%m-%d").date() if args.end else None

    records = pull_new_records(
        start_date=start,
        end_date=end,
        first_pull_days=args.first_pull_days,
        headless=not args.headed,
        all_violations=args.all_violations,
    )
    print(f"New records: {len(records)}")
    for r in records[:3]:
        print(f"  {r.address} | {r.source_url} | {r.raw_text.splitlines()[2]}")


if __name__ == "__main__":
    _cli()
